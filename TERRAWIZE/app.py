import os
import re
import json
import logging
import asyncio
import openpyxl
import sys
import traceback
from dotenv import load_dotenv

# This import statement uses a specialized library that provides "OpenAI" as shown.
# If you normally do "import openai", replace this accordingly.
from openai import OpenAI

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

class ExpertosyRecommendationEngine:
    """
    This class encapsulates all functionality for generating factors, categories,
    factor options, questionnaires, and so forth related to recommending a product
    based on user preferences through the command-line interface.
    """
    def __init__(self, search_query: str):
        """
        Initialize the recommendation engine with a specific search query.
        This query describes what the user is looking for, such as "recommend a power bank."
        """
        self.search_query = search_query
        self.results = {}

    async def generate_factors(self, number_of_factors: int = 10) -> list:
        """
        Generate a set of comprehensive factors for evaluating the item in question.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at analyzing {self.search_query}."
            },
            {
                "role": "user",
                "content": f"List {number_of_factors} unique and comprehensive factors for evaluating {self.search_query}."
            },
            {
                "role": "assistant",
                "content": "Provide factors separated by '*', without numbering or additional text."
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=500,
            messages=messages
        )
        raw_text = response.choices[0].message.content.strip()
        self.results["factors"] = raw_text.split('*')
        return self.results["factors"]

    async def generate_categories(self) -> list:
        """
        Generate a list of categories for the item in question.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at categorizing {self.search_query}."
            },
            {
                "role": "user",
                "content": f"List 50 different categories of {self.search_query}."
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=500,
            messages=messages
        )
        raw_text = response.choices[0].message.content.strip()
        self.results["categories"] = raw_text.split('\n')
        return self.results["categories"]

    async def create_fact_sheet_template(self, factors: list, categories: list) -> str:
        """
        Create a comprehensive fact sheet template for the item in question,
        based on the previously generated factors and categories.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at creating fact sheets for {self.search_query}."
            },
            {
                "role": "user",
                "content": (
                    f"Create a comprehensive fact sheet template for {self.search_query} using these factors: "
                    f"{', '.join(factors)} and categories: {', '.join(categories)}."
                )
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=1000,
            messages=messages
        )
        fact_sheet_template = response.choices[0].message.content.strip()
        self.results["fact_sheet_template"] = fact_sheet_template
        return fact_sheet_template

    async def generate_factor_options(self, factors: list) -> str:
        """
        Generate a set of detailed options for each factor, including a scoring system.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at creating nuanced options for {self.search_query}."
            },
            {
                "role": "user",
                "content": (
                    f"For each factor in {', '.join(factors)}, create 4 detailed options with a scoring system from 0-20."
                )
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=1000,
            messages=messages
        )
        factor_options_text = response.choices[0].message.content.strip()
        self.results["factor_options"] = factor_options_text
        return factor_options_text

    async def generate_famous_models(self, categories: list) -> str:
        """
        Generate a list of famous models or examples of the item in question, categorized accordingly.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at identifying {self.search_query} models."
            },
            {
                "role": "user",
                "content": (
                    f"Create a list of 100 famous {self.search_query} models, categorized by {', '.join(categories)}."
                )
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=1500,
            messages=messages
        )
        famous_models_text = response.choices[0].message.content.strip()
        self.results["famous_models"] = famous_models_text
        return famous_models_text

    async def generate_factor_importance(self, factors: list) -> str:
        """
        Assign importance scores to the previously generated factors.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at evaluating {self.search_query} factors."
            },
            {
                "role": "user",
                "content": (
                    f"Assign importance scores (1-20) for each factor in {', '.join(factors)}."
                )
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=500,
            messages=messages
        )
        factor_importance_text = response.choices[0].message.content.strip()
        self.results["factor_importance"] = factor_importance_text
        return factor_importance_text

    async def create_questionnaire(self, factors: list, factor_options_text: str) -> str:
        """
        Create a comprehensive questionnaire that incorporates a detailed cost breakdown
        for various power bank factors. This method embeds the user-provided cost data
        in the prompt so that the model will generate a multiple-choice set of questions
        each referencing its corresponding costs.
        """
        cost_breakdown_text = """
Here is a breakdown of approximate cost ranges for each option along with justifications based on capacity, features, and market trends. Prices can vary based on brand, retailer, and additional features.

### 1. Capacity (mAh)
- Option A: 10,000 mAh - Cost: $20-$30
- Option B: 20,000 mAh - Cost: $30-$50
- Option C: 30,000 mAh - Cost: $50-$80
- Option D: 50,000 mAh - Cost: $80-$120

### 2. Output Power (W)
- Option A: 5W - Cost: $10-$20
- Option B: 10W - Cost: $15-$30
- Option C: 18W - Cost: $30-$50
- Option D: 30W - Cost: $50-$100

### 3. Number of Ports
- Option A: 1 port - Cost: $10-$20
- Option B: 2 ports - Cost: $20-$30
- Option C: 3 ports - Cost: $30-$50
- Option D: 4+ ports - Cost: $50-$80

### 4. Charging Speed (Fast Charge support)
- Option A: No Fast Charge - Cost: $10-$20
- Option B: Basic Fast Charge (up to 18W) - Cost: $20-$30
- Option C: Quick Charge 3.0 - Cost: $30-$50
- Option D: Power Delivery (PD) - Cost: $50-$100

### 5. Build Quality and Durability
- Option A: Plastic casing - Cost: $10-$20
- Option B: Aluminum casing - Cost: $20-$40
- Option C: Rugged design (shockproof) - Cost: $40-$70
- Option D: Waterproof and shockproof - Cost: $70-$120

### 6. Brand Reputation and Reviews
- Option A: Unknown brand, low reviews - Cost: $10-$20
- Option B: Established brand, mixed reviews - Cost: $20-$30
- Option C: Well-known brand with good reviews - Cost: $40-$70
- Option D: Top-tier brand with excellent reviews - Cost: $70-$120

### 7. Safety Features (overcharge, short circuit protection)
- Option A: No safety features - Cost: $5-$15
- Option B: Basic overcharge protection - Cost: $10-$20
- Option C: Multiple safety features - Cost: $20-$40
- Option D: Advanced safety features - Cost: $40-$70
"""

        messages = [
            {
                "role": "system",
                "content": (
                    f"You are an expert at creating questionnaires for {self.search_query}. "
                    "You have cost breakdown data for each factor (capacity, output power, number of ports, "
                    "charging speed, build quality, brand reputation, and safety features). "
                    "Include these cost details in each multiple-choice option so the user can see how each choice "
                    "affects cost."
                )
            },
            {
                "role": "user",
                "content": (
                    f"{cost_breakdown_text}\n\n"
                    "Now create a single, coherent questionnaire with Q: lines for each major factor. "
                    "Under each question, provide multiple-choice answers labeled '1.', '2.', '3.', '4.', and so on. "
                    "In each answer, reference the relevant cost range from the above breakdown. "
                    "Do not ask the user for anything beyond choosing an option."
                )
            }
        ]

        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=1500,
            messages=messages
        )
        questionnaire_text = response.choices[0].message.content.strip()
        self.results["questionnaire"] = questionnaire_text
        return questionnaire_text

    async def analyze_models(self, famous_models_text: str, factor_options_text: str) -> str:
        """
        Analyze and categorize famous models of the item in question based on the factor options.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at analyzing {self.search_query} models."
            },
            {
                "role": "user",
                "content": (
                    f"Categorize these models: {famous_models_text} using factor options: {factor_options_text}."
                )
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=1500,
            messages=messages
        )
        model_analysis_text = response.choices[0].message.content.strip()
        self.results["model_analysis"] = model_analysis_text
        return model_analysis_text

    async def generate_cost_estimates(self, factor_options_text: str) -> str:
        """
        Generate cost estimates for different options regarding the item in question.
        """
        messages = [
            {
                "role": "system",
                "content": f"You are an expert at pricing {self.search_query}."
            },
            {
                "role": "user",
                "content": f"Provide cost range estimates for these options: {factor_options_text}."
            }
        ]
        response = await self._create_chat_completion(
            model="gpt-4o-mini",
            maximum_tokens=500,
            messages=messages
        )
        cost_estimates_text = response.choices[0].message.content.strip()
        self.results["cost_estimates"] = cost_estimates_text
        return cost_estimates_text

    async def _create_chat_completion(self, model: str, maximum_tokens: int, messages: list):
        """
        A helper method to create a chat completion with error handling,
        using the client.chat.completions.create approach.
        """
        try:
            result = await asyncio.to_thread(
                client.chat.completions.create,
                model=model,
                max_tokens=maximum_tokens,
                messages=messages,
                temperature=0.7
            )
            return result
        except Exception as exception_data:
            logger.error(f"OpenAI API error: {exception_data}")
            raise

    def save_results_to_files(self) -> None:
        """
        Save the results from this recommendation engine to text files
        and compile them into a single Excel file.
        """
        os.makedirs("results", exist_ok=True)

        for key_name, value_data in self.results.items():
            with open(f"results/{key_name}_results.txt", "w", encoding="utf-8") as file_handle:
                file_handle.write(str(value_data))

        excel_workbook = openpyxl.Workbook()
        for key_name, value_data in self.results.items():
            worksheet = excel_workbook.create_sheet(title=key_name[:31])
            if isinstance(value_data, list):
                for index_number, item_data in enumerate(value_data, start=1):
                    worksheet.cell(row=index_number, column=1, value=item_data)
            elif isinstance(value_data, str):
                worksheet.cell(row=1, column=1, value=value_data)

        default_worksheet = excel_workbook["Sheet"]
        if not default_worksheet._cells:
            excel_workbook.remove(default_worksheet)

        excel_workbook.save("results/comprehensive_results.xlsx")


def parse_questionnaire(questionnaire_text: str) -> list:
    """
    Parse an AI-generated questionnaire into a list of dictionaries.
    Each dictionary has 'question' and 'options' as keys.
    The question is derived from lines beginning with 'Q:',
    and multiple-choice options are derived from lines beginning with digits plus a period.
    """
    questions_list = []
    current_question_text = None
    current_options_list = []

    lines = questionnaire_text.split('\n')
    for line in lines:
        stripped_line = line.strip()
        if not stripped_line:
            continue

        # Detect a new question if line starts with "Q:"
        if stripped_line.lower().startswith("q:"):
            if current_question_text and current_options_list:
                questions_list.append({
                    "question": current_question_text,
                    "options": current_options_list
                })
            current_question_text = stripped_line[2:].strip()
            current_options_list = []
        elif re.match(r'^[0-9]+\.', stripped_line):
            option_text = re.sub(r'^[0-9]+\.\s*', '', stripped_line).strip()
            if option_text:
                current_options_list.append(option_text)

    if current_question_text and current_options_list:
        questions_list.append({
            "question": current_question_text,
            "options": current_options_list
        })

    return questions_list


async def run_interactive_recommendation() -> None:
    """
    The main asynchronous routine for the command-line interface flow:
      1. Generate factors, categories, and factor options.
      2. Generate a questionnaire that embeds cost data.
      3. Present the user with multiple-choice questions in the command line.
      4. Once the user provides answers, generate a personalized recommendation.
      5. Save all results to files.
    """
    if len(sys.argv) < 2:
        print("Usage: python app.py 'recommend a power bank'")
        sys.exit(1)

    search_query = ' '.join(sys.argv[1:])
    try:
        engine = ExpertosyRecommendationEngine(search_query)

        # Step 1: Generate factors, categories, and factor options
        generated_factors = await engine.generate_factors()
        generated_categories = await engine.generate_categories()
        generated_factor_options = await engine.generate_factor_options(generated_factors)

        # Step 2: Generate a questionnaire that references cost data
        questionnaire_text = await engine.create_questionnaire(generated_factors, generated_factor_options)

        print(f"\n--- Expertosy Interactive Recommendation for: {search_query} ---")
        print("\nWelcome! Let us find your perfect match through a detailed questionnaire.")

        # Step 3: Parse the questionnaire and interact with the user
        parsed_questions_list = parse_questionnaire(questionnaire_text)
        user_preferences = {}

        if not parsed_questions_list:
            print("\nNo structured questions were detected. The model might not have followed the requested format.")
            print("Below is the raw questionnaire text for debugging purposes:\n")
            print(questionnaire_text)
        else:
            for question_data in parsed_questions_list:
                print(f"\nQ: {question_data['question']}")
                for index_number, option_text in enumerate(question_data["options"], start=1):
                    print(f"  {index_number}. {option_text}")

                while True:
                    try:
                        user_choice = input(
                            f"\nChoose your preference (1-{len(question_data['options'])}): "
                        ).strip()
                        numeric_choice = int(user_choice)
                        if 1 <= numeric_choice <= len(question_data['options']):
                            user_preferences[question_data['question']] = question_data['options'][numeric_choice - 1]
                            break
                        else:
                            print(f"Invalid choice. Please enter a number between 1 and {len(question_data['options'])}.")
                    except ValueError:
                        print("Please enter a valid number.")

        # Step 4: Generate a personalized recommendation from user choices
        print("\n--- Generating Personalized Recommendations ---")
        preference_text = (
            f"I am looking for a {search_query} with the following preferences:\n"
            + "\n".join([
                f"{factor_text}: {answer_text}"
                for factor_text, answer_text in user_preferences.items()
            ])
        )
        recommendation_messages = [
            {
                "role": "system",
                "content": f"You are an expert at recommending {search_query} based on user preferences."
            },
            {
                "role": "user",
                "content": preference_text
            }
        ]
        recommendation_response = await asyncio.to_thread(
            client.chat.completions.create,
            model="gpt-4o-mini",
            max_tokens=1000,
            messages=recommendation_messages,
            temperature=0.7
        )
        final_recommendation = recommendation_response.choices[0].message.content.strip()
        print("--- Your Personalized Recommendation ---")
        print(final_recommendation)

        # Step 5: Save all results
        engine.results["user_preferences"] = user_preferences
        engine.results["recommendation"] = final_recommendation
        engine.save_results_to_files()

        print("\nFull results have been saved in the 'results' directory.")

    except Exception as exception_data:
        logger.error(f"CLI recommendation error: {exception_data}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    asyncio.run(run_interactive_recommendation())
